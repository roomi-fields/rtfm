"""XLSX parser — index Excel workbooks in full, streaming.

Optional dependency: `openpyxl`. If unavailable, the parser is not
registered (the registry simply returns no parser for `.xlsx`).

Per workbook we emit:

  - 1 overview chunk: file path, list of sheets
  - per sheet: schema chunk (column headers with inferred types)
  - per sheet: data chunks covering **every row** (size-bounded,
    header repeated per chunk for column context)

Previously only the first 6 rows per sheet were indexed; now the whole
workbook is searchable. `read_only=True` + `iter_rows` keeps memory
bounded on massive workbooks. Cell values are kept in full.
"""

import hashlib
from pathlib import Path
from typing import Any, Iterator, Optional

from rtfm.core.models import Chunk
from rtfm.parsers.base import BaseParser, ParserRegistry

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:
    load_workbook = None  # type: ignore

# Rows scanned up-front for type inference only.
TYPE_SAMPLE_ROWS = 50
MAX_CHUNK_CHARS = 4000


def _content_hash(text: str) -> str:
    return hashlib.md5(text.encode()).hexdigest()[:12]


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").replace("\r", " ")


def _format_row(headers: list[str], row: list[Any]) -> str:
    parts = []
    for i, h in enumerate(headers):
        val = _format_cell(row[i]) if i < len(row) else ""
        parts.append(f"{h}={val}")
    for j in range(len(headers), len(row)):
        parts.append(f"col{j}={_format_cell(row[j])}")
    return " | ".join(parts)


def _infer_type(values: list[Any]) -> str:
    seen_int = seen_float = seen_bool = seen_str = seen_dt = 0
    for v in values:
        if v is None or v == "":
            continue
        if isinstance(v, bool):
            seen_bool += 1
        elif isinstance(v, int):
            seen_int += 1
        elif isinstance(v, float):
            seen_float += 1
        elif hasattr(v, "isoformat"):
            seen_dt += 1
        else:
            seen_str += 1
    if seen_str:
        return "text"
    if seen_dt:
        return "datetime"
    if seen_float:
        return "float"
    if seen_int:
        return "int"
    if seen_bool:
        return "bool"
    return "empty"


@ParserRegistry.register
class XLSXParser(BaseParser):
    """Parser for Excel `.xlsx` workbooks."""

    extensions = [".xlsx"]
    name = "xlsx"

    @classmethod
    def can_parse(cls, path: Path) -> bool:
        if load_workbook is None:
            return False
        return super().can_parse(path)

    def parse(
        self,
        path: Path,
        metadata: Optional[dict] = None,
    ) -> Iterator[Chunk]:
        if load_workbook is None:
            return
        metadata = metadata or {}

        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except Exception:
            return

        try:
            sheet_names = wb.sheetnames
            if not sheet_names:
                return

            book_title = metadata.get("title", path.name)
            book_slug = metadata.get("book_slug", self._path_to_slug(path))
            book_file = metadata.get("source_file", str(path))
            ext_meta = metadata.get("extended", {})
            idx = 0

            def _new_chunk(chapter_title: str, content: str,
                           row_start: int = 1, row_end: int = 1) -> Chunk:
                nonlocal idx
                idx += 1
                return Chunk(
                    id=f"{book_slug}-{idx:04d}",
                    content=content,
                    book_title=book_title,
                    book_slug=book_slug,
                    book_file=book_file,
                    chapter_title=chapter_title,
                    chapter_num=idx,
                    page_start=row_start,
                    page_end=row_end,
                    paragraph=1,
                    content_chars=len(content),
                    content_hash=_content_hash(content),
                    metadata=ext_meta,
                )

            # Overview first (sheet names only — totals appear per sheet).
            overview = [f"# Workbook: {path.name}", "", "## Sheets"]
            for name in sheet_names:
                overview.append(f"- `{name}`")
            yield _new_chunk("overview", "\n".join(overview))

            # Per sheet: schema chunk, then every row in data chunks.
            for name in sheet_names:
                ws = wb[name]
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    continue
                headers = [str(h) if h is not None else f"col{i+1}"
                           for i, h in enumerate(header_row)]
                if not headers:
                    continue

                # Buffer first rows for type inference, then replay them.
                type_buffer: list[list[Any]] = []
                for row in rows_iter:
                    type_buffer.append(list(row))
                    if len(type_buffer) >= TYPE_SAMPLE_ROWS:
                        break
                types = []
                for col_idx in range(len(headers)):
                    col_values = [r[col_idx] for r in type_buffer if col_idx < len(r)]
                    types.append(_infer_type(col_values))
                schema = [f"# Sheet: {name}", "",
                          f"**{len(headers)} columns**", "", "## Columns"]
                for h, t in zip(headers, types):
                    schema.append(f"- `{h}` *({t})*")
                yield _new_chunk(f"sheet: {name} (schema)", "\n".join(schema))

                # Stream every row into bounded data chunks.
                header_line = f"sheet={name} | columns: " + ", ".join(headers)
                buf: list[str] = []
                buf_chars = 0
                row_start = 1
                row_num = 0
                from itertools import chain
                for row in chain(type_buffer, rows_iter):
                    row_num += 1
                    line = _format_row(headers, list(row))
                    if buf and buf_chars + len(line) + 1 > MAX_CHUNK_CHARS:
                        body = header_line + "\n" + "\n".join(buf)
                        yield _new_chunk(f"sheet: {name} rows {row_start}–{row_num-1}",
                                         body, row_start, row_num - 1)
                        buf = []
                        buf_chars = 0
                        row_start = row_num
                    buf.append(line)
                    buf_chars += len(line) + 1
                if buf:
                    body = header_line + "\n" + "\n".join(buf)
                    yield _new_chunk(f"sheet: {name} rows {row_start}–{row_num}",
                                     body, row_start, row_num)
        finally:
            wb.close()

    def extract_metadata(self, path: Path) -> dict:
        return {
            "source_file": str(path),
            "book_slug": self._path_to_slug(path),
            "title": path.name,
        }

    @staticmethod
    def _path_to_slug(path: Path) -> str:
        return str(path).replace("/", "-").replace("\\", "-").lstrip("-")

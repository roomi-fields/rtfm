"""Tests for rtfm parsers."""

import pytest
from pathlib import Path
from rtfm.parsers.base import ParserRegistry
from rtfm.parsers.xml_legifrance import (
    XMLLegiFranceParser,
    lien_legifrance,
)
from rtfm.parsers.html_bofip import (
    HTMLBOFiPParser,
    extract_cgi_references,
    parse_boi_identifier,
    lien_bofip,
    HTMLTextExtractor,
)
from rtfm.parsers.sqlite_parser import SQLiteParser
from rtfm.parsers.jupyter import JupyterParser
from rtfm.parsers.csv_parser import CSVParser

try:
    from rtfm.parsers.toml_parser import TOMLParser, tomllib as _toml_lib
    _HAS_TOML = _toml_lib is not None
except ImportError:
    TOMLParser = None  # type: ignore
    _HAS_TOML = False

try:
    from rtfm.parsers.xlsx import XLSXParser, load_workbook as _wb
    _HAS_XLSX = _wb is not None
except ImportError:
    XLSXParser = None  # type: ignore
    _HAS_XLSX = False


class TestParserRegistry:
    """Tests for parser registry."""

    def test_registry_has_parsers(self):
        """Test that parsers are registered."""
        extensions = ParserRegistry.list_extensions()
        assert ".xml" in extensions
        assert ".html" in extensions or ".htm" in extensions

    def test_get_parser_for_xml(self):
        """Test getting parser for XML file."""
        parser = ParserRegistry.get_parser(Path("test.xml"))
        assert parser is not None
        assert isinstance(parser, XMLLegiFranceParser)

    def test_get_parser_for_html(self):
        """Test getting parser for HTML file."""
        parser = ParserRegistry.get_parser(Path("test.html"))
        assert parser is not None
        assert isinstance(parser, HTMLBOFiPParser)

    def test_get_parser_unknown(self):
        """Test getting parser for unknown extension."""
        parser = ParserRegistry.get_parser(Path("test.xyz"))
        assert parser is None

    def test_get_parser_for_sqlite(self):
        """Test getting parser for SQLite extensions."""
        for ext in (".sqlite", ".sqlite3"):
            parser = ParserRegistry.get_parser(Path(f"foo{ext}"))
            assert isinstance(parser, SQLiteParser), f"failed for {ext}"


class TestSQLiteParser:
    """Tests for SQLite parser."""

    @pytest.fixture
    def sample_db(self, tmp_path):
        import sqlite3

        db_path = tmp_path / "sample.sqlite"
        conn = sqlite3.connect(db_path)
        conn.executescript(
            """
            CREATE TABLE authors (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL
            );
            CREATE TABLE books (
                id INTEGER PRIMARY KEY,
                title TEXT NOT NULL,
                author_id INTEGER,
                FOREIGN KEY (author_id) REFERENCES authors(id)
            );
            CREATE VIEW recent_books AS SELECT * FROM books ORDER BY id DESC;
            INSERT INTO authors (id, name) VALUES (1, 'Borges'), (2, 'Calvino');
            INSERT INTO books (id, title, author_id) VALUES
                (1, 'Ficciones', 1),
                (2, 'Invisible Cities', 2);
            """
        )
        conn.commit()
        conn.close()
        return db_path

    def test_parse_produces_chunks(self, sample_db):
        parser = SQLiteParser()
        chunks = list(parser.parse(sample_db))
        # overview + (schema + sample) × 2 tables + 1 view = 6
        assert len(chunks) == 6

        titles = [c.chapter_title for c in chunks]
        assert "overview" in titles
        assert "table: authors (schema)" in titles
        assert "table: books (sample)" in titles
        assert "view: recent_books" in titles

    def test_overview_lists_tables_with_counts(self, sample_db):
        parser = SQLiteParser()
        overview = next(c for c in parser.parse(sample_db) if c.chapter_title == "overview")
        assert "authors (2 rows)" in overview.content
        assert "books (2 rows)" in overview.content

    def test_schema_chunk_includes_create_and_columns(self, sample_db):
        parser = SQLiteParser()
        chunk = next(c for c in parser.parse(sample_db) if c.chapter_title == "table: books (schema)")
        assert "CREATE TABLE books" in chunk.content
        assert "`title`" in chunk.content
        assert "`author_id`" in chunk.content
        assert "Foreign Keys" in chunk.content
        assert "authors" in chunk.content

    def test_sample_rows_rendered(self, sample_db):
        parser = SQLiteParser()
        chunk = next(c for c in parser.parse(sample_db) if c.chapter_title == "table: books (sample)")
        assert "Ficciones" in chunk.content
        assert "Invisible Cities" in chunk.content

    def test_extract_edges_returns_foreign_keys(self, sample_db):
        parser = SQLiteParser()
        edges = parser.extract_edges(sample_db)
        assert len(edges) == 1
        assert edges[0].relation_type == "fk"
        assert edges[0].target_ref == "authors"
        assert "books.author_id" in edges[0].source_detail

    def test_db_extension_requires_magic_bytes(self, tmp_path):
        """A `.db` file that is not actually SQLite should be skipped."""
        fake = tmp_path / "fake.db"
        fake.write_text("this is not a sqlite db at all")
        parser = SQLiteParser()
        assert parser.can_parse(fake) is False
        assert list(parser.parse(fake)) == []
        assert parser.extract_edges(fake) == []

    def test_corrupt_db_handled_gracefully(self, tmp_path):
        bad = tmp_path / "bad.sqlite"
        bad.write_bytes(b"\x00" * 100)  # not a valid SQLite file
        parser = SQLiteParser()
        # Doesn't raise — just yields nothing
        assert list(parser.parse(bad)) == []


class TestXMLLegiFranceParser:
    """Tests for Légifrance XML parser."""

    def test_parse_sample_cgi(self, sample_cgi_xml):
        """Test parsing sample CGI XML."""
        parser = XMLLegiFranceParser()
        chunks = list(parser.parse(sample_cgi_xml))

        assert len(chunks) == 2

        # Check first article - chapter_title format is "Article X"
        chunk1 = chunks[0]
        assert "Article" in chunk1.chapter_title
        assert "impôt annuel" in chunk1.content
        assert chunk1.metadata["legal_fr"]["etat"] == "VIGUEUR"

        # Check second article
        chunk2 = chunks[1]
        assert "Article" in chunk2.chapter_title
        assert "exonération" in chunk2.content

    def test_extract_metadata(self, sample_cgi_xml):
        """Test metadata extraction."""
        parser = XMLLegiFranceParser()
        meta = parser.extract_metadata(sample_cgi_xml)

        assert meta["source_file"] == "CGI.xml"
        assert "title" in meta or "book_slug" in meta  # Should have some identification

    def test_lien_legifrance(self):
        """Test Légifrance URL generation."""
        url = lien_legifrance("44 quinquies", "cgi")
        assert "legifrance.gouv.fr" in url
        assert "44" in url or "quinquies" in url


class TestHTMLBOFiPParser:
    """Tests for BOFiP HTML parser."""

    def test_parse_sample_bofip(self, sample_bofip_html):
        """Test parsing sample BOFiP HTML."""
        parser = HTMLBOFiPParser()
        chunks = list(parser.parse(sample_bofip_html))

        # Should have at least one chunk (content is >200 chars)
        assert len(chunks) >= 1

        # Check content contains TVA-related text
        all_content = " ".join(c.content for c in chunks)
        assert "TVA" in all_content or "taxe" in all_content.lower()

    def test_extract_cgi_references(self):
        """Test CGI reference extraction from text."""
        text = "Conformément à l'article 256 du CGI et à l'article 257 bis"
        refs = extract_cgi_references(text)

        assert "256" in refs
        assert "257 bis" in refs

    def test_extract_cgi_refs_with_suffixes(self):
        """Test extraction of articles with Latin suffixes."""
        text = """
        L'article 44 quinquies du CGI prévoit une exonération.
        L'art. 39 decies A permet des amortissements exceptionnels.
        Voir aussi article 238 bis du code général des impôts.
        """
        refs = extract_cgi_references(text)

        assert "44 quinquies" in refs
        assert "39 decies A" in refs
        assert "238 bis" in refs

    def test_extract_cgi_refs_no_false_positives(self):
        """Test that extraction doesn't capture false positives."""
        text = "L'article 44 quinquies du CGI mentionne les zones"
        refs = extract_cgi_references(text)

        # Should not capture "du" or "de" as part of article number
        for ref in refs:
            assert "du" not in ref.lower()
            assert "de" not in ref.lower()

    def test_parse_boi_identifier(self):
        """Test BOI identifier parsing."""
        result = parse_boi_identifier("BOI-TVA-CHAMP-30-10-20-10-20130523")

        assert result["identifiant"] == "BOI-TVA-CHAMP-30-10-20-10-20130523"
        assert result["serie"] == "TVA"
        assert result["division"] == "CHAMP"
        assert result["date_publication"] == "2013-05-23"

    def test_lien_bofip(self):
        """Test BOFiP URL generation."""
        url = lien_bofip("BOI-TVA-CHAMP-30-10-20-10-20130523")
        assert "bofip.impots.gouv.fr" in url
        assert "TVA-CHAMP" in url


class TestHTMLTextExtractor:
    """Tests for HTML text extraction."""

    def test_extract_simple_text(self):
        """Test extracting text from simple HTML."""
        html = "<p>Hello world</p>"
        extractor = HTMLTextExtractor()
        extractor.feed(html)
        text = extractor.get_text()

        assert "Hello world" in text

    def test_extract_with_sections(self):
        """Test extracting sections from HTML."""
        html = """
        <h1>Title</h1>
        <p>Introduction</p>
        <h2>Section 1</h2>
        <p>Content 1</p>
        <h2>Section 2</h2>
        <p>Content 2</p>
        """
        extractor = HTMLTextExtractor()
        extractor.feed(html)
        sections = extractor.get_sections()

        assert len(sections) >= 2
        titles = [s[0] for s in sections]
        assert "Title" in titles or "Section 1" in titles

    def test_skip_script_tags(self):
        """Test that script content is skipped."""
        html = """
        <p>Real content</p>
        <script>var x = 'should not appear';</script>
        <p>More content</p>
        """
        extractor = HTMLTextExtractor()
        extractor.feed(html)
        text = extractor.get_text()

        assert "Real content" in text
        assert "should not appear" not in text

    def test_preserve_list_items(self):
        """Test that list items are preserved."""
        html = """
        <ul>
            <li>Item 1</li>
            <li>Item 2</li>
        </ul>
        """
        extractor = HTMLTextExtractor()
        extractor.feed(html)
        text = extractor.get_text()

        assert "Item 1" in text
        assert "Item 2" in text


class TestJupyterParser:
    """Tests for Jupyter notebook parser."""

    @pytest.fixture
    def sample_notebook(self, tmp_path):
        import json
        nb = {
            "cells": [
                {"cell_type": "markdown", "source": "# Analysis\n\nIntro paragraph."},
                {"cell_type": "code", "source": "import pandas as pd\ndf = pd.read_csv('data.csv')"},
                {"cell_type": "markdown", "source": "## Cleaning"},
                {"cell_type": "code", "source": "df = df.dropna()"},
                {"cell_type": "markdown", "source": "## Modeling"},
                {"cell_type": "code", "source": "from sklearn.linear_model import LogisticRegression"},
            ],
            "metadata": {}, "nbformat": 4, "nbformat_minor": 5,
        }
        path = tmp_path / "demo.ipynb"
        path.write_text(json.dumps(nb))
        return path

    def test_parse_groups_by_heading(self, sample_notebook):
        chunks = list(JupyterParser().parse(sample_notebook))
        titles = [c.chapter_title for c in chunks]
        assert titles == ["Analysis", "Cleaning", "Modeling"]

    def test_code_cells_fenced(self, sample_notebook):
        chunk = next(c for c in JupyterParser().parse(sample_notebook)
                     if c.chapter_title == "Analysis")
        assert "```python" in chunk.content
        assert "import pandas" in chunk.content

    def test_invalid_json_skipped(self, tmp_path):
        bad = tmp_path / "bad.ipynb"
        bad.write_text("not json")
        assert list(JupyterParser().parse(bad)) == []

    def test_registry_routes_ipynb(self):
        parser = ParserRegistry.get_parser(Path("nb.ipynb"))
        assert isinstance(parser, JupyterParser)


@pytest.mark.skipif(not _HAS_TOML, reason="tomllib/tomli not available")
class TestTOMLParser:
    """Tests for TOML parser (pyproject/Cargo style)."""

    @pytest.fixture
    def pyproject(self, tmp_path):
        path = tmp_path / "pyproject.toml"
        path.write_text(
            '[build-system]\n'
            'requires = ["hatchling>=1.0"]\n'
            '\n'
            '[project]\n'
            'name = "demo"\n'
            'version = "0.1.0"\n'
            'dependencies = ["requests>=2.0", "pyyaml"]\n'
            '\n'
            '[project.optional-dependencies]\n'
            'dev = ["pytest>=7.0"]\n'
        )
        return path

    def test_parse_emits_chunks_per_table(self, pyproject):
        chunks = list(TOMLParser().parse(pyproject))
        assert len(chunks) >= 1
        joined = "\n".join(c.chapter_title for c in chunks)
        assert "build-system" in joined
        assert "project" in joined

    def test_extract_dep_edges_pep621(self, pyproject):
        edges = TOMLParser().extract_edges(pyproject)
        targets = {e.target_ref for e in edges}
        assert "requests" in targets
        assert "pyyaml" in targets
        assert "pytest" in targets
        assert "hatchling" in targets
        assert all(e.relation_type == "depends_on" for e in edges)

    def test_extract_dep_edges_cargo(self, tmp_path):
        cargo = tmp_path / "Cargo.toml"
        cargo.write_text(
            '[package]\nname = "demo"\nversion = "0.1.0"\n\n'
            '[dependencies]\nserde = "1.0"\ntokio = { version = "1", features = ["full"] }\n\n'
            '[dev-dependencies]\nproptest = "1.0"\n'
        )
        edges = TOMLParser().extract_edges(cargo)
        targets = {e.target_ref for e in edges}
        assert "serde" in targets
        assert "tokio" in targets
        assert "proptest" in targets

    def test_invalid_toml_skipped(self, tmp_path):
        bad = tmp_path / "bad.toml"
        bad.write_text("this is = not [valid toml ===")
        assert list(TOMLParser().parse(bad)) == []
        assert TOMLParser().extract_edges(bad) == []


class TestCSVParser:
    """Tests for CSV / TSV parser."""

    @pytest.fixture
    def sample_csv(self, tmp_path):
        path = tmp_path / "data.csv"
        path.write_text(
            "name,age,active,score\n"
            "Alice,30,true,98.5\n"
            "Bob,25,false,87.2\n"
            "Carol,42,true,93.1\n"
        )
        return path

    def test_parse_emits_overview_and_sample(self, sample_csv):
        chunks = list(CSVParser().parse(sample_csv))
        titles = [c.chapter_title for c in chunks]
        assert titles == ["overview", "sample"]

    def test_overview_has_column_types(self, sample_csv):
        overview = next(c for c in CSVParser().parse(sample_csv)
                        if c.chapter_title == "overview")
        assert "`name` *(text)*" in overview.content
        assert "`age` *(int)*" in overview.content
        assert "`active` *(bool)*" in overview.content
        assert "`score` *(float)*" in overview.content
        assert "3 data rows" in overview.content

    def test_sample_renders_data(self, sample_csv):
        sample = next(c for c in CSVParser().parse(sample_csv)
                      if c.chapter_title == "sample")
        assert "Alice" in sample.content
        assert "Carol" in sample.content

    def test_tsv_routed_correctly(self, tmp_path):
        path = tmp_path / "data.tsv"
        path.write_text("a\tb\n1\t2\n3\t4\n")
        chunks = list(CSVParser().parse(path))
        assert len(chunks) == 2

    def test_empty_file_yields_nothing(self, tmp_path):
        path = tmp_path / "empty.csv"
        path.write_text("")
        assert list(CSVParser().parse(path)) == []


@pytest.mark.skipif(not _HAS_XLSX, reason="openpyxl not installed")
class TestXLSXParser:
    """Tests for XLSX workbook parser."""

    @pytest.fixture
    def sample_xlsx(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "sales"
        ws1.append(["product", "qty", "price"])
        ws1.append(["Apple", 10, 1.5])
        ws1.append(["Bread", 5, 3.0])
        ws2 = wb.create_sheet("inventory")
        ws2.append(["sku", "stock"])
        ws2.append(["A001", 100])
        path = tmp_path / "wb.xlsx"
        wb.save(path)
        return path

    def test_parse_emits_overview_and_per_sheet(self, sample_xlsx):
        chunks = list(XLSXParser().parse(sample_xlsx))
        titles = [c.chapter_title for c in chunks]
        # 1 overview + (schema + sample) × 2 sheets
        assert "overview" in titles
        assert "sheet: sales (schema)" in titles
        assert "sheet: sales (sample)" in titles
        assert "sheet: inventory (schema)" in titles
        assert "sheet: inventory (sample)" in titles

    def test_overview_lists_sheets_with_dimensions(self, sample_xlsx):
        overview = next(c for c in XLSXParser().parse(sample_xlsx)
                        if c.chapter_title == "overview")
        assert "`sales`" in overview.content
        assert "`inventory`" in overview.content
        assert "2 rows × 3 cols" in overview.content

    def test_sample_includes_data(self, sample_xlsx):
        sample = next(c for c in XLSXParser().parse(sample_xlsx)
                      if c.chapter_title == "sheet: sales (sample)")
        assert "Apple" in sample.content
        assert "Bread" in sample.content

    def test_corrupt_xlsx_handled(self, tmp_path):
        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not a real xlsx")
        # Doesn't raise — yields nothing
        assert list(XLSXParser().parse(bad)) == []
